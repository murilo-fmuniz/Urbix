from fastapi import APIRouter, HTTPException
from pydantic import BaseModel
from typing import List, Dict, Optional
from pathlib import Path
from openpyxl import load_workbook
import logging

# Configurar logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

router = APIRouter()

class Indicator(BaseModel):
    CODRM: Optional[str] = None
    NOME_RM: Optional[str] = None
    ANO: Optional[int] = None
    NOTA_INTELIGENTE: float
    NOTA_SUSTENTAVEL: float
    ESPVIDA: Optional[float] = None
    FECTOT: Optional[float] = None
    IDHM: Optional[float] = None
    IDHM_E: Optional[float] = None
    IDHM_L: Optional[float] = None
    IDHM_R: Optional[float] = None
    T_FLFUND_TUDO: Optional[float] = None
    T_FLMED_TUDO: Optional[float] = None
    T_FLBAS_TUDO: Optional[float] = None
    T_FUND11A13_TUDO: Optional[float] = None
    RAZDEP: Optional[float] = None

def get_column_letter(idx: int) -> str:
    """Converte índice numérico em letra de coluna do Excel (A, B, C, etc)"""
    result = ""
    while idx > 0:
        idx, rem = divmod(idx-1, 26)
        result = chr(65 + rem) + result
    return result

def normalizar(valor: float, minimo: float, maximo: float) -> float:
    """Normaliza um valor entre 0 e 1 usando min-max scaling."""
    if maximo == minimo:
        return 0
    return (valor - minimo) / (maximo - minimo)

def calcular_nota(dados: Dict[str, float], min_max: Dict[str, tuple]) -> float:
    """Calcula a nota de cidade inteligente baseada nos indicadores."""
    nota = 0
    peso_total = 0
    
    # ESPVIDA (Expectativa de Vida)
    if "ESPVIDA" in dados and "ESPVIDA" in min_max and dados["ESPVIDA"] is not None:
        try:
            peso = 0.2
            valor_norm = normalizar(float(dados["ESPVIDA"]), min_max["ESPVIDA"][0], min_max["ESPVIDA"][1])
            nota += valor_norm * peso
            peso_total += peso
        except (ValueError, TypeError) as e:
            logger.warning(f"Erro ao processar ESPVIDA: {e}")

    # IDHM (Índice de Desenvolvimento Humano Municipal)
    if "IDHM" in dados and "IDHM" in min_max and dados["IDHM"] is not None:
        try:
            peso = 0.3
            valor_norm = normalizar(float(dados["IDHM"]), min_max["IDHM"][0], min_max["IDHM"][1])
            nota += valor_norm * peso
            peso_total += peso
        except (ValueError, TypeError) as e:
            logger.warning(f"Erro ao processar IDHM: {e}")

    # FECTOT (Taxa de Fecundidade Total)
    if "FECTOT" in dados and "FECTOT" in min_max and dados["FECTOT"] is not None:
        try:
            peso = 0.1
            valor_norm = normalizar(float(dados["FECTOT"]), min_max["FECTOT"][0], min_max["FECTOT"][1])
            nota += valor_norm * peso
            peso_total += peso
        except (ValueError, TypeError) as e:
            logger.warning(f"Erro ao processar FECTOT: {e}")

    # Normalizar a nota final
    return (nota / peso_total) if peso_total > 0 else 0

def calcular_nota_sustentavel(dados: Dict[str, float], min_max: Dict[str, tuple]) -> float:
    """Calcula a nota de cidade sustentável baseada nos indicadores."""
    nota = 0
    peso_total = 0
    
    # IDHM_R (IDHM Renda)
    if "IDHM_R" in dados and "IDHM_R" in min_max and dados["IDHM_R"] is not None:
        try:
            peso = 0.4
            valor_norm = normalizar(float(dados["IDHM_R"]), min_max["IDHM_R"][0], min_max["IDHM_R"][1])
            nota += valor_norm * peso
            peso_total += peso
        except (ValueError, TypeError) as e:
            logger.warning(f"Erro ao processar IDHM_R: {e}")

    # IDHM_L (IDHM Longevidade)
    if "IDHM_L" in dados and "IDHM_L" in min_max and dados["IDHM_L"] is not None:
        try:
            peso = 0.3
            valor_norm = normalizar(float(dados["IDHM_L"]), min_max["IDHM_L"][0], min_max["IDHM_L"][1])
            nota += valor_norm * peso
            peso_total += peso
        except (ValueError, TypeError) as e:
            logger.warning(f"Erro ao processar IDHM_L: {e}")

    # RAZDEP (Razão de Dependência)
    if "RAZDEP" in dados and "RAZDEP" in min_max and dados["RAZDEP"] is not None:
        try:
            peso = 0.3
            valor_norm = normalizar(float(dados["RAZDEP"]), min_max["RAZDEP"][0], min_max["RAZDEP"][1])
            nota += valor_norm * peso
            peso_total += peso
        except (ValueError, TypeError) as e:
            logger.warning(f"Erro ao processar RAZDEP: {e}")

    # Normalizar a nota final
    return (nota / peso_total) if peso_total > 0 else 0

@router.get("/indicators", response_model=List[Indicator])
async def get_indicators():
    """Retorna todos os indicadores com notas calculadas."""
    file = Path(__file__).parent.parent / 'data' / 'dataBase.xlsx'
    
    try:
        logger.info(f"Abrindo arquivo: {file}")
        wb = load_workbook(filename=str(file), read_only=True, data_only=True)
        ws = wb.active
        
        # Obter cabeçalhos
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        logger.debug(f"Cabeçalhos encontrados: {headers}")
        
        # Preparar dados para cálculo de min/max
        dados_brutos = []
        min_max = {
            "ESPVIDA": [float('inf'), float('-inf')],
            "IDHM": [float('inf'), float('-inf')],
            "FECTOT": [float('inf'), float('-inf')],
            "IDHM_R": [float('inf'), float('-inf')],
            "IDHM_L": [float('inf'), float('-inf')],
            "RAZDEP": [float('inf'), float('-inf')]
        }
        
        # Primeira passagem: calcular min/max
        for row in ws.iter_rows(min_row=2):
            for idx, cell in enumerate(row):
                if idx < len(headers) and headers[idx] in min_max and cell.value is not None:
                    try:
                        valor = float(cell.value)
                        min_max[headers[idx]][0] = min(min_max[headers[idx]][0], valor)
                        min_max[headers[idx]][1] = max(min_max[headers[idx]][1], valor)
                    except (ValueError, TypeError) as e:
                        logger.warning(f"Erro ao converter valor para float: {cell.value} na coluna {headers[idx]}")
        
        # Segunda passagem: ler dados
        for row in ws.iter_rows(min_row=2):
            registro = {}
            try:
                for idx, cell in enumerate(row):
                    if idx < len(headers):  # Garantir que não ultrapasse o número de cabeçalhos
                        if cell.value is not None:
                            if headers[idx] == "CODRM":
                                registro[headers[idx]] = str(cell.value)
                            elif headers[idx] in ["ESPVIDA", "IDHM", "IDHM_E", "IDHM_L", "IDHM_R", "FECTOT", "RAZDEP"]:
                                try:
                                    registro[headers[idx]] = float(cell.value)
                                except (ValueError, TypeError):
                                    registro[headers[idx]] = None
                            elif headers[idx] == "ANO":
                                try:
                                    registro[headers[idx]] = int(cell.value)
                                except (ValueError, TypeError):
                                    registro[headers[idx]] = None
                            else:
                                registro[headers[idx]] = cell.value
                        else:
                            registro[headers[idx]] = None
                dados_brutos.append(registro)
            except Exception as e:
                logger.error(f"Erro ao processar linha: {e}")
                continue
        
        logger.debug(f"Min/Max calculados: {min_max}")
        
        # Calcular notas
        resultados = []
        for dados in dados_brutos:
            try:
                notas = {}
                notas["NOTA_INTELIGENTE"] = calcular_nota(dados, min_max)
                notas["NOTA_SUSTENTAVEL"] = calcular_nota_sustentavel(dados, min_max)
                dados_completos = {**dados, **notas}
                resultados.append(Indicator(**dados_completos))
            except Exception as e:
                logger.error(f"Erro ao calcular notas: {e}")
                logger.debug(f"Dados problemáticos: {dados}")
                continue
        
        wb.close()
        return resultados
        
    except FileNotFoundError:
        raise HTTPException(
            status_code=404,
            detail="Arquivo dataBase.xlsx não encontrado"
        )
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Erro ao processar dados: {str(e)}"
        )

@router.get("/indicators/summary")
async def get_indicators_summary():
    """Retorna um resumo estatístico dos indicadores."""
    file = Path(__file__).parent.parent / 'data' / 'dataBase.xlsx'
    
    try:
        wb = load_workbook(filename=str(file), read_only=True, data_only=True)
        ws = wb.active
        
        # Obter cabeçalhos
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        
        # Inicializar estruturas para estatísticas
        stats = {header: {'min': float('inf'), 'max': float('-inf'), 'sum': 0, 'count': 0} 
                for header in headers if header}
        
        # Calcular estatísticas
        for row in ws.iter_rows(min_row=2):
            for idx, cell in enumerate(row):
                if cell.value is not None and headers[idx] in stats:
                    try:
                        valor = float(cell.value)
                        stats[headers[idx]]['min'] = min(stats[headers[idx]]['min'], valor)
                        stats[headers[idx]]['max'] = max(stats[headers[idx]]['max'], valor)
                        stats[headers[idx]]['sum'] += valor
                        stats[headers[idx]]['count'] += 1
                    except (ValueError, TypeError):
                        continue
        
        # Calcular médias e formatar resultado
        resumo = {}
        for header, valores in stats.items():
            if valores['count'] > 0:
                resumo[header] = {
                    'min': valores['min'],
                    'max': valores['max'],
                    'mean': valores['sum'] / valores['count'],
                    'count': valores['count']
                }
        
        wb.close()
        return resumo
        
    except FileNotFoundError:
        raise HTTPException(
            status_code=404,
            detail="Arquivo dataBase.xlsx não encontrado"
        )
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Erro ao processar dados: {str(e)}"
        )
